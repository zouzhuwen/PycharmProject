import pyexcel_xls
import openpyxl


def read_data(path,sheet):
    """读取excel文件"""
    #获取path路径下的xlsx文件为data 字典
    data = pyexcel_xls.get_data(path)
    # for list in data.keys():#  list为xlsx文件内存在的表名
        # if str(list)==sheet:
        #     for l in data[list]: #  data为sheet表格的内容，每一行为一个列表
        #         print(l)        #  打印每一行
    return data[sheet]          #  返回表格名为sheet的表格内容

def write_data(path,data,sheetName):
    try:
         """写入excel文件
         path:文件路径
         data:写入的数据
         sheetName：表格名称
         """
         workbook = openpyxl.Workbook(path) #  创建一个excel对象 path为文件路径
         sheet = workbook.create_sheet(sheetName)  #  创建一个名为sheetName的表格
         for i in range(len(data)):
             sheet.append(data[i])
         workbook.save(filename=path)
         print("保存{}表格成功，路径：{}".format(sheetName,path))
    except Exception as e:
        print("写入{}文件失败：{}".format(sheetName,e))


def add_sheet(path,data,newsheet):
    try:
        """增加一个sheet表格
        path:文件路径
        data:写入的数据
        newsheet:需要增加的表格
        """
        workbook = openpyxl.load_workbook(path)
        sheet = workbook.create_sheet(newsheet)
        for i in range(len(data)):
            sheet.append(data[i])
        workbook.save(path)
        print("新增{}表格成功，路径{}".format(newsheet,path))
    except Exception as e:
        print("新增{}表格失败:{}".format(newsheet,e))



if __name__ == '__main__':
    data = read_data(r'C:\Users\hp\Desktop\工作簿1.xlsx','Sheet1')
    # print(data)

    write_data(r'C:\Users\hp\Desktop\工作簿2.xlsx',data,"sheet2")

    add_sheet(r'C:\Users\hp\Desktop\工作簿2.xlsx',data,'newsheet3')
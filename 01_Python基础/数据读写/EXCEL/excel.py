# -*- coding: utf-8 -*-
# @Author: 罗湘飞
# @Date  : 2019/2/23/023

import pyexcel_xls
import xlsxwriter
import openpyxl
from log import logging_test
class Excel(object):
    '''excel文件操作类'''
    def __init__(self):
        # self.log=logger
        pass
    def get_xls(self,url,sheet):
        """读取ecxel文件"""
        self.data_xls=pyexcel_xls.get_data(url)
        #这里读取excel文件，
        for get_sheet in self.data_xls.keys():
            if str(get_sheet)==sheet:
                return self.data_xls[get_sheet]
    def write_xls(self,listNum,xlsname="测试结果.xlsx",sheet="sheet"):
        '''使用xlsxwriter方法来实现excel数据保存，给变量数据listNum（二维列表），xlsname为保存excel绝对路径'''
        try:
            self.workbook = xlsxwriter.Workbook (xlsname)  #创建一个excel文件
            self.worksheet = self.workbook.add_worksheet (sheet)    #在文件中创建一个名为TEST的sheet,不加名字默认为sheet1
            for row_num, row_data in enumerate (listNum):   #循环读取excel文件
                self.worksheet.write_row (row_num, 0, row_data)  #把每一行再写入到excel中 从第一行第一个单元格开始写
            self.workbook.close ()
            print("用例执行并保存成功")
        except Exception as e:
            print("用例保存失败,失败原因：{}".format(e))
    def write_newxls(self,data,xlsxname,sheet="测试结果"):
        '''使用openpyxl方法来实现excel数据保存，给变量数据data（二维列表），xlsname为保存excel绝对路径'''
        try:
            workbook = openpyxl.Workbook (xlsxname)
            ws1 = workbook.create_sheet (sheet)
            for i in range (len (data)):
                ws1.append (data[i])
            workbook.save (filename=xlsxname)
            print("用例执行并保存为{}成功".format(xlsxname))
        except Exception as e:
            print("用例保存失败,失败原因：{}".format(e))
    def amend_xls(self,data,xlsxname,sheet="测试结果"):
        '''使用openpyxl方法来实现在原有excel文件上增加sheet'''
        try:
            workbook = openpyxl.load_workbook (xlsxname)
            ws1 = workbook.create_sheet (sheet)  # 创建一个sheet
            for i in range (len (data)):
                ws1.append (data[i])
            workbook.save (filename=xlsxname)
            print("用例执行并保存为{}成功,sheet为{}".format(xlsxname,sheet))
        except Exception as e:
            print("用例保存失败,失败原因：{}".format(e))
if __name__ == '__main__':
    # a=Excel()
    # b=a.get_xls(r"E:\python_project\K8s_port_test\output_file/20190513-212808测试用例结果.xlsx","call_log_data2")
    # for i in b:
    #     i.append
    e = Excel()
    data = e.get_xls(r'C:\Users\hp\Desktop\工作簿1.xlsx','Sheet1')
    print(data)
    for list in data:
        list.append("yes")
        print(list)

    e.write_xls(data,xlsname=r'C:\Users\hp\Desktop\测试结果.xlsx',sheet='表格')

    e.write_newxls(data,r'C:\Users\hp\Desktop\测试结果_new.xlsx',sheet='表格')
    e.amend_xls(data,r'C:\Users\hp\Desktop\测试结果_new.xlsx',sheet='new_sheet')